"""
Phase 6 – Price List Matching
=============================

This module:
- Loads the price list from a local .xlsx file (LOCAL_PRICE_LIST_PATH) or
  from a Google Sheet (fallback).
- Matches normalized part names to price list entries.
- Returns PART_NUMBER, PRICE; flags unmatched rows (never drops them).

Matching priority (as per spec):
1. Exact PART_NUMBER match (if part_number already present in input)
2. Exact canonical name match (normalized, case-insensitive)
3. AI-powered matching via Gemini (for abbreviated/handwritten OCR text)
4. Fuzzy match (>=65% similarity) as fallback

Guardrails:
- Must only be used when config.ENABLE_ORDER_UPLOAD is True.
- Does NOT modify existing GST parsing or invoice logic.
"""
from __future__ import annotations

from typing import List, Dict, Optional, Tuple
import os
import re
import json

try:
    import config
except ImportError:
    pass

try:
    from src import config
except ImportError:
    pass


def _normalize_for_matching(text: str) -> str:
    """
    Normalize part name for fuzzy matching:
    lowercase, collapse spaces, remove punctuation.
    """
    s = (text or "").lower().strip()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _fuzzy_similarity(a: str, b: str) -> float:
    """
    Simple token-based fuzzy similarity (0.0 to 1.0).
    
    For now, we use a simple ratio of common tokens.
    For production, consider difflib.SequenceMatcher or fuzzywuzzy.
    """
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a, b).ratio()


class PriceListLoader:
    """Load price list from .xlsx or Google Sheets."""

    @staticmethod
    def load_from_xlsx(path: str) -> List[Dict]:
        """
        Load price list from .xlsx file.
        
        Expected columns:
        - PART_NUMBER
        - PART_NAME_CANONICAL (or PART_NAME, DESCRIPTION, etc.)
        - PRICE
        - ALIASES (optional, comma-separated)
        
        Returns list of dicts with normalized keys.
        """
        if not os.path.exists(path):
            return []
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[PriceListLoader] openpyxl not installed; cannot load .xlsx price list")
            return []
        
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            return []
        
        # Try to find header row
        header = None
        data_start = 1
        for idx, row in enumerate(rows_raw):
            if row and any(
                str(c or "").upper().replace(" ", "").replace(".", "") in [
                    "PARTNO", "PARTNUMBER", "PARTNUM",
                    "DESCRIPTION", "PARTNAME",
                    "PRICE", "MRP"
                ]
                for c in row
            ):
                header = [str(c or "").strip() for c in row]
                data_start = idx + 1
                break
        
        if not header:
            # No recognizable header; assume first row is header
            header = [str(c or "").strip() for c in rows_raw[0]]
            data_start = 1
        
        # Normalize header keys
        def norm_key(k: str) -> str:
            k_upper = k.upper().replace(" ", "_")
            # Handle "Part No." or "PART_NUMBER"
            if ("PART" in k_upper and "NO" in k_upper) or (
                "PART" in k_upper and "NUM" in k_upper
            ):
                return "PART_NUMBER"
            # Handle "Description" or "PART_NAME"
            if "DESC" in k_upper or (
                "PART" in k_upper and "NAME" in k_upper
            ):
                return "PART_NAME_CANONICAL"
            # Handle "MRP" or "PRICE"
            if "PRICE" in k_upper or "MRP" in k_upper:
                return "PRICE"
            if "ALIAS" in k_upper:
                return "ALIASES"
            return k
        
        header_map = {idx: norm_key(h) for idx, h in enumerate(header)}
        
        # Handle merged cells: if a "PRICE" key exists and the next column has None header,
        # check if data exists there and map it to PRICE
        for idx in range(len(header)):
            if header_map.get(idx) == "PRICE":
                # Check if next column(s) have None but contain numeric data
                for offset in range(1, 3):
                    next_idx = idx + offset
                    if next_idx < len(header) and (
                        not header[next_idx] or str(header[next_idx]).strip().upper() == "NONE"
                    ):
                        # Check if this column has numeric data in sample rows
                        has_numeric = False
                        for sample_row in rows_raw[data_start:data_start + 10]:
                            if next_idx < len(sample_row):
                                val = sample_row[next_idx]
                                if val is not None and str(val).strip():
                                    try:
                                        float(str(val).strip())
                                        has_numeric = True
                                        break
                                    except (ValueError, TypeError):
                                        pass
                        if has_numeric:
                            header_map[next_idx] = "PRICE"
                            break
        
        result = []
        for row in rows_raw[data_start:]:
            if not row or not any(row):
                continue
            d = {}
            for idx, val in enumerate(row):
                key = header_map.get(idx)
                if key:
                    d[key] = str(val or "").strip()
            if d.get("PART_NUMBER") or d.get("PART_NAME_CANONICAL") or d.get("PRICE"):
                result.append(d)
        
        return result

    @staticmethod
    def _parse_sheet_rows(rows: List[List[str]]) -> List[Dict]:
        """
        Parse raw sheet rows (header + data) into normalised price list dicts.

        Handles two layouts:
        - Clean:  Part No. | Description | MRP | STD PKG | MASTER PKG
        - Merged: Part No. | Description | MRP (long header) | (blank col with price) | STD PKG | MASTER PKG
        """
        if not rows or len(rows) < 2:
            return []

        header = rows[0]

        # Detect column indices dynamically
        pn_idx = desc_idx = price_idx = std_idx = master_idx = None
        for i, h in enumerate(header):
            h_upper = (h or "").upper().replace(" ", "")
            if "PARTNO" in h_upper or "PARTNUM" in h_upper:
                pn_idx = i
            elif "DESC" in h_upper or "PARTNAME" in h_upper:
                desc_idx = i
            elif "MRP" in h_upper or "PRICE" in h_upper:
                price_idx = i
            elif "STDPKG" in h_upper:
                std_idx = i
            elif "MASTERPKG" in h_upper:
                master_idx = i

        if pn_idx is None or desc_idx is None:
            return []

        # If the price column header is the merged "MRP (Incl. ...)" type,
        # actual price data may be in the NEXT column.  Detect by sampling.
        real_price_idx = price_idx
        if price_idx is not None:
            # check if the identified column is actually empty and the next has numbers
            sample_has_data = False
            sample_next_has_data = False
            for r in rows[1:min(20, len(rows))]:
                val = r[price_idx] if price_idx < len(r) else ""
                if val and val.strip():
                    try:
                        float(val.strip().replace(",", ""))
                        sample_has_data = True
                    except (ValueError, TypeError):
                        pass
                nxt = price_idx + 1
                if nxt < len(r):
                    nval = r[nxt] if nxt < len(r) else ""
                    if nval and nval.strip():
                        try:
                            float(nval.strip().replace(",", ""))
                            sample_next_has_data = True
                        except (ValueError, TypeError):
                            pass
            if not sample_has_data and sample_next_has_data:
                real_price_idx = price_idx + 1

        result = []
        for row in rows[1:]:
            pn = (row[pn_idx] if pn_idx < len(row) else "").strip()
            desc = (row[desc_idx] if desc_idx < len(row) else "").strip()
            price = ""
            if real_price_idx is not None and real_price_idx < len(row):
                price = (row[real_price_idx] or "").strip().replace(",", "")
            std_pkg = ""
            if std_idx is not None and std_idx < len(row):
                std_pkg = (row[std_idx] or "").strip()
            master_pkg = ""
            if master_idx is not None and master_idx < len(row):
                master_pkg = (row[master_idx] or "").strip()

            if pn or desc or price:
                result.append({
                    "PART_NUMBER": pn,
                    "PART_NAME_CANONICAL": desc,
                    "PRICE": price,
                    "STD_PKG": std_pkg,
                    "MASTER_PKG": master_pkg,
                })
        return result

    @staticmethod
    def load_from_sheets(spreadsheet) -> List[Dict]:
        """
        Load and merge price list from Google Sheets tabs:
          1. Price_List2  (clean, primary)
          2. Price_List1  (fallback for items not in #1)
          3. Correct standard packaging  (overrides/corrections – highest priority)

        Merge strategy:
          - Keyed by PART_NUMBER (uppercase).
          - Price_List2 entries are loaded first.
          - Price_List1 fills gaps (items not already present).
          - Correct standard packaging overwrites any existing entry.

        Results are cached to a local JSON file to avoid repeated API reads.
        Cache is refreshed if older than 60 minutes.
        """
        import json
        import time

        # Check for local cache first
        cache_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "temp")
        os.makedirs(cache_dir, exist_ok=True)
        cache_path = os.path.join(cache_dir, "price_list_cache.json")
        CACHE_TTL = 3600  # 60 minutes

        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    cache = json.load(f)
                if time.time() - cache.get("timestamp", 0) < CACHE_TTL:
                    items = cache.get("items", [])
                    print(f"[PriceListLoader] Loaded {len(items)} items from cache (age: {int(time.time() - cache['timestamp'])}s)")
                    return items
                else:
                    print("[PriceListLoader] Cache expired, reloading from Sheets...")
            except Exception:
                pass  # corrupt cache, reload

        # Load fresh from Sheets
        merged: dict = {}  # keyed by PART_NUMBER (upper)

        tab_order = [
            ("Price_List2", False),                     # primary
            ("Price_List1", False),                     # fill gaps
            ("Correct standard packaging", True),       # override
        ]

        for tab_name, is_override in tab_order:
            try:
                ws = spreadsheet.worksheet(tab_name)
                rows = ws.get_all_values()
                items = PriceListLoader._parse_sheet_rows(rows)
                added = 0
                for item in items:
                    pn = (item.get("PART_NUMBER") or "").strip().upper()
                    if not pn:
                        continue
                    if is_override or pn not in merged:
                        merged[pn] = item
                        added += 1
                print(f"[PriceListLoader] {tab_name}: {len(items)} rows parsed, {added} {'overridden' if is_override else 'added'}")
            except Exception as e:
                print(f"[PriceListLoader] Could not load tab '{tab_name}': {e}")

        result = list(merged.values())
        print(f"[PriceListLoader] Total merged price list: {len(result)} unique items")

        # Save to cache
        try:
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump({"timestamp": time.time(), "items": result}, f)
            print(f"[PriceListLoader] Cache saved to {cache_path}")
        except Exception as e:
            print(f"[PriceListLoader] Failed to save cache: {e}")

        return result


class PriceListMatcher:
    """
    Match normalized part names to price list entries.
    
    Matching priority:
    1. Exact PART_NUMBER (if already set)
    2. Exact canonical name (case-insensitive, normalized)
    3. Fuzzy match on canonical or aliases (>=90%)
    """

    def __init__(self, price_list: List[Dict]):
        self.price_list = price_list or []
        self._by_part_number: Dict[str, Dict] = {}
        self._by_canonical: Dict[str, Dict] = {}
        self._aliases_map: Dict[str, Dict] = {}

        for entry in self.price_list:
            pn = (entry.get("PART_NUMBER") or "").strip().upper()
            canonical = _normalize_for_matching(entry.get("PART_NAME_CANONICAL") or "")
            price_s = entry.get("PRICE") or ""
            try:
                price = float(price_s)
            except (ValueError, TypeError):
                price = None

            record = {
                "PART_NUMBER": pn,
                "PART_NAME_CANONICAL": entry.get("PART_NAME_CANONICAL") or "",
                "PRICE": price,
            }

            if pn:
                self._by_part_number[pn] = record
            if canonical:
                self._by_canonical[canonical] = record

            aliases_raw = entry.get("ALIASES") or ""
            for alias in aliases_raw.replace(",", " ").split():
                a_norm = _normalize_for_matching(alias)
                if a_norm and a_norm not in self._aliases_map:
                    self._aliases_map[a_norm] = record

    def match(
        self,
        part_name: str,
        part_number: Optional[str] = None,
        min_score: float = 0.65,
    ) -> Tuple[Optional[str], Optional[float], str]:
        """
        Match a single part name (and optional part_number) to the price list.
        
        Returns:
            (PART_NUMBER, PRICE, match_type)
            - match_type: "EXACT_PN", "EXACT_NAME", "FUZZY", "UNMATCHED"
        """
        # 1) Exact PART_NUMBER
        if part_number:
            pn_upper = part_number.strip().upper()
            if pn_upper in self._by_part_number:
                rec = self._by_part_number[pn_upper]
                return (rec["PART_NUMBER"], rec["PRICE"], "EXACT_PN")

        norm = _normalize_for_matching(part_name)
        if not norm:
            return (None, None, "UNMATCHED")

        # 2) Exact canonical match
        if norm in self._by_canonical:
            rec = self._by_canonical[norm]
            return (rec["PART_NUMBER"], rec["PRICE"], "EXACT_NAME")

        # 3) Exact alias match
        if norm in self._aliases_map:
            rec = self._aliases_map[norm]
            return (rec["PART_NUMBER"], rec["PRICE"], "EXACT_NAME")

        # 4) Token-overlap pre-filter + fuzzy match
        input_tokens = set(norm.split())
        best_score = 0.0
        best_rec = None

        candidates = []
        for canon, rec in self._by_canonical.items():
            canon_tokens = set(canon.split())
            common = input_tokens & canon_tokens
            if len(common) >= 2 or (input_tokens and len(common) / len(input_tokens) >= 0.5):
                candidates.append((canon, rec))

        if not candidates:
            candidates = list(self._by_canonical.items())[:500]

        for canon, rec in candidates:
            score = _fuzzy_similarity(norm, canon)
            if score > best_score:
                best_score = score
                best_rec = rec

        if best_score >= min_score and best_rec:
            return (best_rec["PART_NUMBER"], best_rec["PRICE"], "FUZZY")

        return (None, None, "UNMATCHED")

    def match_batch_with_ai(
        self,
        ocr_names: List[str],
    ) -> Dict[str, Tuple[Optional[str], Optional[float], str]]:
        """
        Use Gemini AI to match a batch of OCR-extracted (abbreviated/handwritten)
        part names to canonical price list entries.

        This is far more accurate than fuzzy string matching because the LLM
        understands abbreviations, misspellings, and shorthand.

        Returns:
            Dict mapping each OCR name -> (PART_NUMBER, PRICE, match_type)
        """
        import google.generativeai as genai

        results: Dict[str, Tuple[Optional[str], Optional[float], str]] = {}

        # First, do exact/fuzzy matches for easy ones
        ai_needed = []
        for name in ocr_names:
            pn, price, mtype = self.match(name)
            if mtype != "UNMATCHED":
                results[name] = (pn, price, mtype)
            else:
                ai_needed.append(name)

        if not ai_needed:
            return results

        # Build a condensed price list for the prompt (part_number -> description)
        # Only include items with part numbers and prices
        price_entries = []
        for entry in self.price_list:
            pn = (entry.get("PART_NUMBER") or "").strip()
            desc = (entry.get("PART_NAME_CANONICAL") or "").strip()
            price = (entry.get("PRICE") or "").strip()
            if pn and desc and price:
                price_entries.append(f"{pn} | {desc} | {price}")

        # Chunk the price list to fit in context (send ~2000 entries at a time)
        CHUNK_SIZE = 2000
        price_chunks = [price_entries[i:i+CHUNK_SIZE] for i in range(0, len(price_entries), CHUNK_SIZE)]

        prompt = f"""You are matching handwritten order items to a product price list.

The handwritten items use abbreviations and shorthand. For example:
- "HFDLS" = "Head Light" or "Headlight"
- "suspoid" = "Suspension" 
- "Pass pro" = "Passion Pro"
- "BL/wrey" = "Blue/Grey" or "Black/Grey"
- "Bodey Kit" = "Body Kit"
- "visor" = "Visor" or "Nose" (front visor)
- "SP" = "Splendor"
- "BSG" = "BS6"
- "i3s" = "I3S"
- "Type S" or "Type" = model variant
- "m/Grey" = "Matt Grey"
- "S/Red" = "Silver/Red" or "Sports Red"
- "Bkh" = "Black"
- "Bluak" = "Black"
- "Blarek" = "Black"
- "orrenye" = "Orange"
- "wree" = part of color description
- "Access" = "Access" (Suzuki model)
- "Dream Neo" = "Dream Neo" (Honda model)
- "Duet" = "Duet" (TVS model)
- "Xpro" = "X-Pro" or "XPro"
- "old" = older model variant

Here are the handwritten items to match:
{json.dumps(ai_needed, indent=2)}

Below is the product price list (Part Number | Description | Price):
{chr(10).join(price_entries[:CHUNK_SIZE])}

For EACH handwritten item, find the BEST matching product from the price list.
Consider the vehicle model name, part type, and color when matching.

Respond in STRICT JSON format only — no markdown, no explanation:
{{
  "matches": [
    {{"ocr_name": "the handwritten text", "part_number": "SAI-XXX", "confidence": "high/medium/low"}},
    ...
  ]
}}

If no reasonable match exists, set part_number to null.
Include ALL {len(ai_needed)} items in your response.
"""

        try:
            genai.configure(api_key=config.GOOGLE_API_KEY)
            model = genai.GenerativeModel("gemini-2.5-flash")
            response = model.generate_content(prompt)
            response_text = (response.text or "").strip()

            # Clean markdown fences if present
            if response_text.startswith("```"):
                response_text = re.sub(r"^```\w*\n?", "", response_text)
                response_text = re.sub(r"\n?```$", "", response_text)

            ai_result = json.loads(response_text)
            matches = ai_result.get("matches", [])

            for m in matches:
                ocr_name = m.get("ocr_name", "")
                pn = m.get("part_number")
                confidence = m.get("confidence", "low")

                if not pn or not ocr_name:
                    if ocr_name:
                        results[ocr_name] = (None, None, "UNMATCHED")
                    continue

                pn_upper = pn.strip().upper()
                if pn_upper in self._by_part_number:
                    rec = self._by_part_number[pn_upper]
                    results[ocr_name] = (rec["PART_NUMBER"], rec["PRICE"], f"AI_{confidence.upper()}")
                else:
                    # Part number from AI not in our list
                    results[ocr_name] = (None, None, "UNMATCHED")

            # Fill in any items not returned by AI
            for name in ai_needed:
                if name not in results:
                    results[name] = (None, None, "UNMATCHED")

            matched_by_ai = sum(1 for v in results.values() if v[2].startswith("AI_"))
            print(f"[PriceListMatcher] AI matched {matched_by_ai}/{len(ai_needed)} items")

        except Exception as e:
            print(f"[PriceListMatcher] AI matching failed: {e}")
            # Fall back to UNMATCHED for all AI-needed items
            for name in ai_needed:
                if name not in results:
                    results[name] = (None, None, "UNMATCHED")

        return results
